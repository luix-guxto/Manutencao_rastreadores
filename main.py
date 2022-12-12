import datetime

import openpyxl as xl

# cria a pasta ./result caso nao exista
import os
if not os.path.exists('result'):
    os.makedirs('result')

# Cria o objeto de planilha
ativosPlanilha = xl.Workbook()
engespPlanilha = xl.Workbook()

if not os.path.exists('planilhas'):
    os.makedirs('planilhas')
    print("Coloque as planilhas \"ativos_rastreador.xlsx\" e \"base_engesp.xlsx\" na pasta planilhas e execute "
          "novamente o programa \n\n- pressione enter para sair -")
    input()
    exit(1)

if not os.path.exists('planilhas/ativos_rastreador.xlsx'):
    print("Coloque a planilha \"ativos_rastreador.xlsx\" na pasta planilhas e execute novamente o programa \n\n- "
          "pressione enter para sair -")
    input()
    exit(1)
if not os.path.exists('planilhas/base_engesp.xlsx'):
    print("Coloque a planilha \"base_engesp.xlsx\" na pasta planilhas e execute novamente o programa \n\n- pressione "
          "enter para sair -")
    input()
    exit(1)

# carrega as duas planilhas
try:
    ativosPlanilha = xl.load_workbook('./planilhas/ativos_rastreador.xlsx')
except FileNotFoundError:
    print('Arquivo "./planilhas/ativos_rastreador.xlsx" nao encontrado')
    exit()

try:
    engespPlanilha = xl.load_workbook('./planilhas/base_engesp.xlsx')
except FileNotFoundError:
    print('Arquivo "./planilhas/base_engesp.xlsx" nao encontrado')
    exit()

# seleciona a unica aba das planilhas
ativosAba = ativosPlanilha.active
engespAba = engespPlanilha.active

# verifica se as planilhas estão vazias
if ativosAba.max_row == 1:
    print('Planilha "ativos_rastreador.xlsx" vazia')
    exit()
if engespAba.max_row == 1:
    print('Planilha "engesp_rastreador.xlsx" vazia')
    exit()

# coloca todas as placas da planilha de ativos em uma lista (segunda coluna)
placasAtivos = []
for linha in ativosAba.iter_rows(min_row=2, min_col=2, max_col=2):
    for celula in linha:
        placasAtivos.append(celula.value.replace('-', '').upper())

# coloca todos os dados da planilha de engesp em uma lista (todas as colunas)
dadosEngesp = []
for linha in engespAba.iter_rows(min_row=2, min_col=1, max_col=8):
    linhaData = [celula.value for celula in linha]
    linhaData[1] = linhaData[1].replace('-', '').upper()
    dadosEngesp.append(linhaData)

# cria uma nova planilha onde serao colocados os dados da engesp das placas (segunda coluna) não estão na placasAtivos
planilhaRetirada = xl.Workbook()
abaRetirada = planilhaRetirada.active

# cria uma nova planilha para as placas que mudaram o nome do proprietario
planilhaMudanca = xl.Workbook()
abaMudanca = planilhaMudanca.active

# cria um cabecalho para a planilha nova [nome, placa, grupo]
abaRetirada.append(['Nome', 'Placa', 'Grupo'])
abaMudanca.append(['NomeAntigo', 'NomeNovo', 'Placa', 'Telefone'])

for linha in dadosEngesp:
    if linha[1] not in placasAtivos:
        abaRetirada.append(linha)

    nome = linha[0].upper()
    nome = nome.replace('Ç', 'C')
    nomeDiferente = ""
    try:
        nomeDiferente = ativosAba.cell(row=placasAtivos.index(linha[1])+2, column=1).value
    except ValueError:
        pass
    nomeDiferente = nomeDiferente.upper()
    nomeDiferente = nomeDiferente.replace('Ç', 'C')

    if linha[1] in placasAtivos and nome != nomeDiferente:
        # telefone coluna 3 da aba de ativos
        telefone = ativosAba.cell(row=placasAtivos.index(linha[1])+2, column=3).value
        linhaMudanca = [nome, nomeDiferente, linha[1], telefone]
        abaMudanca.append(linhaMudanca)

# calcula o numero total de placas da nova planilha e o valor total da coluna 3
totalLinhas = abaRetirada.max_row - 1
totalValor = 0
totalLinhasMudanca = abaMudanca.max_row - 1
for linha in abaRetirada.iter_rows(min_row=2, min_col=3, max_col=3):
    for celula in linha:
        totalValor += celula.value

# cria um novo arquivo txt com a data de hoje informando o valor total e o numero de linhas
with open(f'./result/total_{datetime.date.today()}.txt', 'w') as arquivo:
    arquivo.write(f'Valor total da mensalidade a ser retirado: R$ {totalValor},00\n\n')
    arquivo.write(f'total de placas a ser retiradas: {totalLinhas}\n')
    arquivo.write(f'total de placas que mudaram o nome do proprietario: {totalLinhasMudanca}')

# salva a nova planilha
planilhaRetirada.save('./result/engesp_retirada.xlsx')
planilhaMudanca.save('./result/engesp_mudanca.xlsx')

# fecha as planilhas
ativosPlanilha.close()
engespPlanilha.close()
planilhaRetirada.close()

# crated by @luix-guxto 2022
